from functools import wraps
from flask import Flask, request, jsonify, make_response
from flask_cors import CORS
from unidecode import unidecode
from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.utils import FORMULAE
import math  # Importa el módulo math para usar math.isnan()
import requests
import pandas as pd
import csv
import json
import os
import sys
import io
import pika
from pika.exchange_type import ExchangeType
import base64
import threading


app = Flask(__name__)
CORS(app)

# url = os.environ.get('APP_URL')
rabbit_url = os.environ.get('RABBITMQ_URL')
rabbit_queue = os.environ.get("RABBITMQ_QUEUE")
rabbit_python_queue = os.environ.get("RABBITMQ_PYTHON_QUEUE")


def main():
    connection = pika.BlockingConnection(
        pika.ConnectionParameters(host=rabbit_url))
    channel = connection.channel()

    channel.queue_declare(queue=rabbit_python_queue, durable=True)

    def callback(ch, method, properties, body):
        mensaje_decodificado = body.decode('utf-8')

        # Analiza el mensaje como un objeto JSON
        mensaje_json = json.loads(mensaje_decodificado)

        # Accede a los datos dentro del objeto 'data'
        datos = mensaje_json['data']

        # Accede a elementos específicos dentro de 'data'
        nombre_proveedor = datos['nombreProveedor']
        base64 = datos['base64']

        print(" [x] Received:")
        print("Nombre del proveedor:", nombre_proveedor)
        print("Base64:", base64)

    channel.basic_consume(
        queue=rabbit_python_queue, on_message_callback=callback, auto_ack=True)

    print(' [*] Waiting for messages. To exit press CTRL+C')
    channel.start_consuming()


def enviar_resultado_a_rabbitmq(channel, nombre_proveedor, data):
    mensaje = {
        "pattern": "carga_tabla",
        "data": {
            "proveedor_actualizado": nombre_proveedor,
            "resultado": data,
        }
    }

    mensaje_json = json.dumps(mensaje)

    # channel.exchange_declare(exchange='api_nest', exchange_type='direct')

    channel.basic_publish(
        exchange='',
        routing_key=rabbit_queue,
        body=mensaje_json,
        properties=pika.BasicProperties(
            delivery_mode=2,  # Hace que el mensaje sea persistente
        )
    )

    return "Resultado enviado a RabbitMQ"


# Direccion archivos
listadosTemporales = "archivosPorCargar/temporales/"
listadoCsv = "archivosPorCargar/csv/"
listadoJson = "archivosPorCargar/archivosJson/"
diccionarios = "nuevosScripts/diccionarios/diccionarios.json"

# Función para encontrar un valor en el diccionario


@app.route('/procesar_archivo', methods=['POST'])
def procesar_archivo():
    try:
        # Obtener los datos JSON del cuerpo de la solicitud
        data = request.get_json()

        # Verificar si 'nombreProveedor' y 'base64' están en el JSON
        if 'nombreProveedor' not in data or 'base64' not in data:
            return jsonify({"error": "El JSON debe contener 'nombreProveedor' y 'base64'"}), 400

        nombre_proveedor = data['nombreProveedor']
        archivo_base64 = data['base64']

        # Llamar a la función de procesamiento adecuada según el proveedor
        if nombre_proveedor == 'air':
            resultado = procesar_archivo_air(archivo_base64)
        elif nombre_proveedor == 'eikon':
            resultado = procesar_archivo_eikon(archivo_base64)
        elif nombre_proveedor == 'elit':
            resultado = procesar_archivo_elit(archivo_base64)
        elif nombre_proveedor == 'hdc':
            resultado = procesar_archivo_hdc(archivo_base64)
        elif nombre_proveedor == 'invid':
            resultado = procesar_archivo_invid(archivo_base64)
        elif nombre_proveedor == 'nb':
            resultado = procesar_archivo_nb(archivo_base64)
        elif nombre_proveedor == 'mega':
            resultado = procesar_archivo_mega(archivo_base64)
        else:
            return jsonify({"error": "Proveedor no válido"}), 400

        connection = pika.BlockingConnection(
            pika.ConnectionParameters(rabbit_url))
        channel = connection.channel()
        respuesta = enviar_resultado_a_rabbitmq(
            channel, nombre_proveedor, resultado)
        connection.close()
        return jsonify({"respuesta": respuesta})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def encontrar_valor(diccionario, clave):
    if clave in diccionario:
        return diccionario[clave]
    else:
        if (clave == "ESTABILIZADORES - UPS - Zapatillas Eléctricas"):
            return diccionario["ESTABILIZADORES - UPS - Zapatillas Electricas"]
        elif (clave == "GPS - De Exploración"):
            return diccionario["GPS - De Exploracion"]
        elif (clave == "TV - Iluminación"):
            return diccionario["TV - Iluminacion"]
        else:
            return "Varios"

# Función para obtener un diccionario


def obtenerDiccionario(nombreDiccionario):
    with open(diccionarios) as diccionariosOpen:
        diccionariosJson = json.load(diccionariosOpen)
    diccionarioBuscado = diccionariosJson[nombreDiccionario]
    return diccionarioBuscado


def tablaAir(archivo_bytesios):
    csv_reader = csv.reader(io.TextIOWrapper(archivo_bytesios), delimiter=',')

    # Crea una lista para almacenar los datos
    data = []
    next(csv_reader)  # Ignora la primera fila de encabezados
    for row in csv_reader:
        if (row[2] != "A"):
            descripcion = row[1]
            rubro = row[10]
            iva = row[4]
            precio = row[2]

            # Crea un diccionario con los datos de cada registro
            registro = {
                'proveedor': 'air',
                'producto': descripcion,
                'categoria': encontrar_valor(obtenerDiccionario('air'), rubro),
                'precio': round((float(precio) * (1 + (float(iva)/100)) * 1.1))
            }

            # Agrega el diccionario a la lista de datos
            data.append(registro)

    return data


def procesar_archivo_air(archivo_base64):

    file_data = base64.b64decode(archivo_base64)
    archivo_bytesio = io.BytesIO(file_data)

    data = tablaAir(archivo_bytesio)

    return data


def tablaEikon(archivo_bytesio):
    df = pd.read_excel(archivo_bytesio)

    # Se borran las primeras filas
    df = df.drop([0, 1, 2])
    df.reset_index(drop=True, inplace=True)

    # Procesar los datos del DataFrame df
    data = []

    for index, row in df.iterrows():
        descripcion = row[1]
        categoria = row[6]
        precio = row[3]

        registro = {
            "proveedor": "eikon",
            "producto": descripcion,
            "categoria": encontrar_valor(obtenerDiccionario("eikon"), categoria),
            "precio": round((float(precio) * 1.1))
        }

        data.append(registro)

    return data


def procesar_archivo_eikon(archivo_base64):

    file_data = base64.b64decode(archivo_base64)
    archivo_bytesio = io.BytesIO(file_data)

    data = tablaEikon(archivo_bytesio)

    return data


def tablaElit(archivo_bytesio):
    df = pd.read_excel(archivo_bytesio)

    data = []

    for index, row in df.iterrows():
        descripcion = row[1]
        categoria = row[5]
        precio = row[7]
        iva = row[8]
        ivaInterno = row[9]

        registro = {
            "proveedor": "elit",
            "producto": descripcion,
            "categoria": encontrar_valor(obtenerDiccionario("elit"), categoria),
            "precio": round((float(precio) * (1 + (float(iva) + float(ivaInterno)) / 100) * 1.1))
        }

        data.append(registro)

    return data


def procesar_archivo_elit(archivo_base64):

    file_data = base64.b64decode(archivo_base64)
    archivo_bytesio = io.BytesIO(file_data)

    data = tablaElit(archivo_bytesio)

    return data


def obtenerTipoIva(clave):
    # Diccionario con tipo de IVA
    tipoIva = {
        "002-I.V.A. 10.5 %": 10.5,
        "001-I.V.A. 21 %": 21,
        "005-Impuestos Internos": 21
    }
    if clave in tipoIva:
        return tipoIva[clave]


def tablaHdc(archivo_bytesio):
    df = pd.read_excel(archivo_bytesio)

    # Procesar los datos en memoria
    data = []

    for index, row in df.iterrows():
        if not math.isnan(row[7]):  # Verificar si no es NaN
            descripcion = row[5]
            if row[3]:
                categoria = unidecode(str(row[3]))
            else:
                categoria = unidecode(str(row[2]))
            precio = row[7]
            iva = obtenerTipoIva(row[8])

            # Crea un diccionario con los datos de cada registro
            registro = {
                "proveedor": "hdc",
                "producto": descripcion,
                "categoria": encontrar_valor(obtenerDiccionario("hdc"), categoria),
                "precio": round(float(precio) * (1+float(iva)/100) * 1.1)
            }

            # Agrega el diccionario a la lista de datos
            data.append(registro)

    return data


def procesar_archivo_hdc(archivo_base64):

    file_data = base64.b64decode(archivo_base64)
    archivo_bytesio = io.BytesIO(file_data)
    data = tablaHdc(archivo_bytesio)

    return data


def tablaInvid(archivo_bytesio):
    df = pd.read_excel(archivo_bytesio)

    df = df.drop([0, 1, 2, 3, 4, 5, 6, 7, 8])
    df.reset_index(drop=True, inplace=True)

    wb = Workbook()
    ws = wb.active

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Aplicar la lógica para calcular la columna "categoria"
    ws['I3'] = "categoria"

    def apply_custom_formula(value1, value2, value3, value4, value5):
        result = ""
        if value1 is None or (isinstance(value1, str) and len(value1) <= 1):
            result = ""
        elif (value2 is None or (isinstance(value2, str) and len(value2) <= 1)) and (value3 is None or (isinstance(value3, str) and len(value3) <= 1)):
            result = value4
        else:
            result = value5
        return result

    for row in range(4, ws.max_row + 1):
        value1 = ws[f'H{row}'].value
        value2 = ws[f'A{row-2}'].value
        value3 = ws[f'C{row-2}'].value
        value4 = ws[f'B{row-2}'].value
        value5 = ws[f'I{row-1}'].value
        result = apply_custom_formula(
            value1, value2, value3, value4, value5)

        cell = ws.cell(row=row, column=9)
        cell.value = unidecode(result)

    # Procesar los datos en memoria sin necesidad de guardar como CSV
    data = []

    for row in range(4, ws.max_row + 1):
        if (ws.cell(row=row, column=4).value and
                ws.cell(row=row, column=4).value != "Nro. de Parte" and
                str(ws.cell(row=row, column=4).value).isdigit()):  # Verificar si es un número válido
            descripcion = ws.cell(row=row, column=2).value
            categoria = unidecode(ws.cell(row=row, column=9).value)
            precio = float(ws.cell(row=row, column=6).value)
            iva = (1 + (float(ws.cell(row=row, column=7).value)/100)) * \
                (1 + (float(ws.cell(row=row, column=8).value)/100))

            registro = {
                "proveedor": "invid",
                "producto": descripcion,
                "categoria": encontrar_valor(obtenerDiccionario("invid"), categoria),
                "precio": round((precio * iva * 1.1))
            }

            data.append(registro)

    return data


def procesar_archivo_invid(archivo_base64):

    file_data = base64.b64decode(archivo_base64)
    archivo_bytesio = io.BytesIO(file_data)
    data = tablaInvid(archivo_bytesio)

    return data


def tablaNb(archivo_bytesio):
    csv_data = archivo_bytesio.read().decode('utf-8').splitlines()

    # Crear una lista para almacenar los datos
    data = []

    # Iterar sobre las filas del archivo CSV (ignorando la primera fila de encabezados)
    csv_reader = csv.reader(csv_data, delimiter=";")
    next(csv_reader)  # Ignorar la primera fila de encabezados

    for row in csv_reader:
        descripcion = row[3]
        categoria = row[2]
        precio = row[10]

        # Crear un diccionario con los datos de cada registro
        registro = {
            "proveedor": "nb",
            "producto": descripcion,
            "categoria": encontrar_valor(obtenerDiccionario("nb"), categoria),
            "precio": round((float(precio) * 1.1))
        }

        # Agregar el diccionario a la lista de datos
        data.append(registro)

    return data


def procesar_archivo_nb(archivo_base64):

    file_data = base64.b64decode(archivo_base64)
    archivo_bytesio = io.BytesIO(file_data)
    data = tablaNb(archivo_bytesio)

    return data


def tablaMega(archivo_bytesio):
    df = pd.read_excel(archivo_bytesio)

    # Se utilizan los índices 0, 1 y 2 para las primeras tres filas
    df = df.drop([0, 1])
    df.reset_index(drop=True, inplace=True)
    df.to_excel(listadosTemporales + "tempMega.xlsx", index=False)

    # Cargar el libro de trabajo
    book = load_workbook(listadosTemporales + "tempMega.xlsx")
    sheet = book["Sheet1"]

    def apply_custom_formula(value1, value2, value3, value4, value5, value6, value7):
        result = ""
        if (value1 is None or len(value1) <= 1) and (value2 is None or len(value2) <= 1):
            result = f"{value3} - {value4}"
        elif (value5 is None or len(value5) <= 1):
            result = ""
        elif (value2 is None or len(value2) <= 1):
            if (value6 is not None):
                posicion = value6.find(" - ")
                if posicion != -1:
                    parte_izquierda = value6[:posicion]
                    result = f"{parte_izquierda} - {value4}"
        else:
            result = f"{value7}"
        return result

    for row in range(3, sheet.max_row + 1):
        value1 = sheet[f'B{row-2}'].value
        value2 = sheet[f'B{row-1}'].value
        value3 = sheet[f'A{row-2}'].value
        value4 = sheet[f'A{row-1}'].value
        value5 = sheet[f'E{row}'].value
        value6 = sheet[f'G{row-2}'].value
        value7 = sheet[f'G{row-1}'].value
        result = apply_custom_formula(
            value1, value2, value3, value4, value5, value6, value7)

        cell = sheet.cell(row=row, column=7)  # Columna G
        cell.value = unidecode(result)

    # Save the changes to the file
    book.save(listadosTemporales + "tempMega.xlsx")

    # Guardar como CSV
    df = pd.read_excel(listadosTemporales+"tempMega.xlsx")
    df.to_csv(listadoCsv+"listadoMega.csv", index=False)

    # Abre el archivo CSV en modo lectura con la codificación adecuada
    with open(listadoCsv+"listadoMega.csv", "r", encoding="utf-8") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=",")

        # Crea una lista para almacenar los datos
        data = []

        # Lee cada fila del archivo CSV (ignorando la primera fila de encabezados)

        next(csv_reader)  # Ignora la primera fila de encabezados
        for row in csv_reader:
            if (row[3] != ""):
                descripcion = row[1]
                precio = row[2].replace("U$s ", "")
                iva = row[4].replace("+", "").replace("%", "")
                categoria = row[6]

                # Crea un diccionario con los datos de cada registro
                registro = {
                    "proveedor": "mega",
                    "producto": descripcion,
                    "categoria": encontrar_valor(obtenerDiccionario("mega"), categoria),
                    "precio": round((float(precio) * (1 + (float(iva)/100)) * 1.1))
                }

                # Agrega el diccionario a la lista de datos

                data.append(registro)

    return data


def procesar_archivo_mega(archivo_base64):

    file_data = base64.b64decode(archivo_base64)
    archivo_bytesio = io.BytesIO(file_data)
    data = tablaMega(archivo_bytesio)

    return data


if __name__ == "__main__":
    # certfile = 'secrets/fullchain.pem'
    # keyfile = 'secrets/privkey.pem'
    rabbitmq_thread = threading.Thread(target=main)
    rabbitmq_thread.start()

    app.run(host='0.0.0.0', debug=True
            # , ssl_context=(certfile, keyfile)
            )

# if __name__ == '__main__':
#     try:
#         main()
#     except KeyboardInterrupt:
#         print('Interrupted')
#         try:
#             sys.exit(0)
#         except SystemExit:
#             os._exit(0)
