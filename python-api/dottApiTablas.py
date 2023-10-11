from functools import wraps
from flask import Flask, request, jsonify, make_response
from flask_cors import CORS
from unidecode import unidecode
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.utils import FORMULAE
import requests
import pandas as pd
import csv
import json
import os
import io
import pika


app = Flask(__name__)
CORS(app)  # This will enable CORS for all routes

url = os.environ.get('APP_URL')
rabbit_url = os.environ.get('RABBITMQ_URL')
rabbit_queue = os.environ.get("RABBITMQ_QUEUE")


def enviar_resultado_a_rabbitmq(channel, data):
    # Prepara el mensaje a enviar (proveedor actualizado y JSON de resultado)
    mensaje = {
        "proveedor_actualizado": "air",
        "resultado": data
    }

    # Convierte el mensaje a JSON
    mensaje_json = json.dumps(mensaje)

    # Envía el mensaje a la cola
    channel.basic_publish(
        exchange='',
        routing_key=rabbit_queue,
        body=mensaje_json,
        properties=pika.BasicProperties(
            delivery_mode=2,  # Hace que el mensaje sea persistente
        )
    )

    return "Resultado enviado a RabbitMQ"


# Direccion archivos
listadosTemporales = "archivosPorCargar/temporales/"
listadoCsv = "archivosPorCargar/csv/"
listadoJson = "archivosPorCargar/archivosJson/"
diccionarios = "nuevosScripts/diccionarios/diccionarios.json"

# Función para encontrar un valor en el diccionario


def encontrar_valor(diccionario, clave):
    if clave in diccionario:
        return diccionario[clave]
    else:
        if (clave == "ESTABILIZADORES - UPS - Zapatillas Eléctricas"):
            return diccionario["ESTABILIZADORES - UPS - Zapatillas Electricas"]
        elif (clave == "GPS - De Exploración"):
            return diccionario["GPS - De Exploracion"]
        elif (clave == "TV - Iluminación"):
            return diccionario["TV - Iluminacion"]
        else:
            return "Varios"

# Función para obtener un diccionario


def obtenerDiccionario(nombreDiccionario):
    with open(diccionarios) as diccionariosOpen:
        diccionariosJson = json.load(diccionariosOpen)
    diccionarioBuscado = diccionariosJson[nombreDiccionario]
    return diccionarioBuscado


# Decorador personalizado para verificar el token
def requiere_token(f):
    @wraps(f)
    def decorador(*args, **kwargs):
        token = request.headers.get("Authorization")
        if not token:
            # Código de respuesta 401 significa no autorizado
            return jsonify({"mensaje": "Acceso no autorizado"}), 401
        # Pasamos el token como primer argumento al método de vista
        return f(token, *args, **kwargs)
    return decorador


def tablaAir():
    with open(listadosTemporales+"temp_air.csv", 'r', encoding="iso-8859-1") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=',')

        # Crea una lista para almacenar los datos
        data = []
        next(csv_reader)  # Ignora la primera fila de encabezados
        for row in csv_reader:
            if (row[2] != "A"):
                descripcion = row[1]
                rubro = row[10]
                iva = row[4]
                precio = row[2]

                # Crea un diccionario con los datos de cada registro
                registro = {
                    'proveedor': 'air',
                    'producto': descripcion,
                    'categoria': encontrar_valor(obtenerDiccionario('air'), rubro),
                    'precio': round((float(precio) * (1 + (float(iva)/100)) * 1.1))
                }

                # Agrega el diccionario a la lista de datos
                data.append(registro)

    return data


@app.route("/procesar_archivo_air", methods=["POST"])
@requiere_token
def procesar_archivo_air(token):

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    archivo.save(listadosTemporales+"temp_air.csv")

    data = tablaAir()

    headers = {
        "Authorization": f"{token}"
    }

    # requests.post(url=url, json=data, headers=headers,
    #               timeout=None, verify=False)

    return make_response(jsonify(data), 200)


def tablaEikon():
    df = pd.read_excel(listadosTemporales+"temp_eikon.xlsx")

    # Se borran las primeras filas
    df = df.drop([0, 1, 2])
    df.reset_index(drop=True, inplace=True)

    # Guardar como CSV
    df.to_csv(listadoCsv+"listadoEik.csv", index=False)

    with open(listadoCsv+"listadoEik.csv", "r") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=",")

        # Crea una lista para almacenar los datos
        data = []

        next(csv_reader)  # Ignora la primera fila de encabezados
        for row in csv_reader:

            descripcion = row[1]
            categoria = row[6]
            precio = row[3]

            # Crea un diccionario con los datos de cada registro
            registro = {
                "proveedor": "eikon",
                "producto": descripcion,
                "categoria": encontrar_valor(obtenerDiccionario("eikon"), categoria),
                "precio": round((float(precio) * 1.1))
            }

            # Agrega el diccionario a la lista de datos
            data.append(registro)

    with open(listadoJson+"eik.json", 'w') as jf:
        json.dump(data, jf, ensure_ascii=False, indent=2)

    return data


@app.route("/procesar_archivo_eikon", methods=["POST"])
@requiere_token
def procesar_archivo_eikon(token):

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    archivo.save(listadosTemporales+"temp_eikon.xlsx")

    data = tablaEikon()

    headers = {
        "Authorization": f"{token}"
    }

    # response = requests.post(
    #     url=url, json=data, headers=headers, timeout=None, verify=False)

    return make_response(jsonify(data), 200)


def tablaElit():
    df = pd.read_excel(listadosTemporales+"temp_elit.xlsx")
    df.to_csv(listadoCsv+"listadoElit.csv", index=False)

    # Abre el archivo CSV en modo lectura con la codificación adecuada
    with open(listadoCsv+"listadoElit.csv", "r") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=",")

        # Crea una lista para almacenar los datos
        data = []

        next(csv_reader)  # Ignora la primera fila de encabezados
        for row in csv_reader:
            descripcion = row[1]
            categoria = row[5]
            precio = row[7]
            iva = row[8]
            ivaInterno = row[9]

            # Crea un diccionario con los datos de cada registro
            registro = {
                "proveedor": "elit",
                "producto": descripcion,
                "categoria": encontrar_valor(obtenerDiccionario("elit"), categoria),
                "precio": round((float(precio) * (1 + (float(iva) + float(ivaInterno))/100) * 1.1))
            }

            # Agrega el diccionario a la lista de datos
            data.append(registro)

    with open(listadoJson+"elit.json", 'w') as jf:
        json.dump(data, jf, ensure_ascii=False, indent=2)

    return data


@app.route("/procesar_archivo_elit", methods=["POST"])
@requiere_token
def procesar_archivo_elit(token):

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    archivo.save(listadosTemporales+"temp_elit.xlsx")

    data = tablaElit()

    headers = {
        "Authorization": f"{token}"
    }

    # requests.post(url=url, json=data, headers=headers,
    #               timeout=None, verify=False)

    return make_response(jsonify(data), 200)


def obtenerTipoIva(clave):
    # Diccionario con tipo de IVA
    tipoIva = {
        "002-I.V.A. 10.5 %": 10.5,
        "001-I.V.A. 21 %": 21,
        "005-Impuestos Internos": 21
    }
    if clave in tipoIva:
        return tipoIva[clave]


def tablaHdc():
    df = pd.read_excel(listadosTemporales+"temp_hdc.xlsx")
    df.to_csv(listadoCsv+"listadoHdc.csv", index=False)

    # Abre el archivo CSV en modo lectura con la codificación adecuada
    with open(listadoCsv+"listadoHdc.csv", "r", encoding="utf-8") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=",")

        # Crea una lista para almacenar los datos
        data = []

        next(csv_reader)  # Ignora la primera fila de encabezados

        for row in csv_reader:
            if row[7]:
                descripcion = row[5]
                if row[3]:
                    categoria = unidecode(row[3])
                else:
                    categoria = unidecode(row[2])
                precio = row[7]
                iva = obtenerTipoIva(row[8])

                # Crea un diccionario con los datos de cada registro
                registro = {
                    "proveedor": "hdc",
                    "producto": descripcion,
                    "categoria": encontrar_valor(obtenerDiccionario("hdc"), categoria),
                    "precio": round(float(precio) * (1+float(iva)/100) * 1.1)
                }

                # Agrega el diccionario a la lista de datos
                data.append(registro)

    with open(listadoJson+"hdc.json", 'w') as jf:
        json.dump(data, jf, ensure_ascii=False, indent=2)

    return data


@app.route("/procesar_archivo_hdc", methods=["POST"])
@requiere_token
def procesar_archivo_hdc(token):

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    archivo.save(listadosTemporales+"temp_hdc.xlsx")

    data = tablaHdc()

    headers = {
        "Authorization": f"{token}"
    }

    # requests.post(url=url, json=data, headers=headers,
    #               timeout=None, verify=False)

    return make_response(jsonify(data), 200)


def tablaInvid(archivo):
    # Guardar el archivo .xls en el directorio
    df = pd.read_excel(archivo)

    # Se utilizan los índices 0, 1 y 2 para las primeras tres filas
    df = df.drop([0, 1, 2, 3, 4, 5, 6, 7, 8])
    df.reset_index(drop=True, inplace=True)
    df.to_excel(listadosTemporales + "temp_invid.xlsx", index=False)

    # Cargar el libro de trabajo
    book = load_workbook(listadosTemporales + "temp_invid.xlsx")
    sheet = book["Sheet1"]

    sheet["I3"] = "categoria"

    def apply_custom_formula(value1, value2, value3, value4, value5):
        result = ""
        if (value1 is None or len(value1) <= 1):
            result = ""
        elif (value2 is None or len(value2) <= 1) and (value3 is None or len(value3) <= 1):
            result = value4
        else:
            result = value5
        return result

    for row in range(4, sheet.max_row + 1):
        value1 = sheet[f'H{row}'].value
        value2 = sheet[f'A{row-2}'].value
        value3 = sheet[f'C{row-2}'].value
        value4 = sheet[f'B{row-2}'].value
        value5 = sheet[f'I{row-1}'].value
        result = apply_custom_formula(
            value1, value2, value3, value4, value5)

        cell = sheet.cell(row=row, column=9)
        cell.value = unidecode(result)

    # Save the changes to the file
    book.save(listadosTemporales + "temp_invid.xlsx")

    # Guardar como CSV
    df = pd.read_excel(listadosTemporales+"temp_invid.xlsx")
    df.to_csv(listadoCsv+"listadoInvid.csv", index=False)

    with open(listadoCsv+"listadoInvid.csv", "r") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=",")

        # Crea una lista para almacenar los datos
        data = []

        # Lee cada fila del archivo CSV (ignorando la primera fila de encabezados)

        next(csv_reader)  # Ignora la primera fila de encabezados
        for row in csv_reader:
            if (row[3] != "" and row[3] != "Nro. de Parte"):
                descripcion = row[1]
                categoria = unidecode(row[8])
                precio = float(row[5])
                iva = (1 + (float(row[6])/100)) * (1 + (float(row[7])/100))

                # Crea un diccionario con los datos de cada registro
                registro = {
                    "proveedor": "invid",
                    "producto": descripcion,
                    "categoria": encontrar_valor(obtenerDiccionario("invid"), categoria),
                    "precio": round((precio * iva * 1.1))
                }

                # Agrega el diccionario a la lista de datos

                data.append(registro)

    with open(listadoJson+"invid.json", 'w') as jf:
        json.dump(data, jf, ensure_ascii=False, indent=2)

    return data


@app.route("/procesar_archivo_invid", methods=["POST"])
@requiere_token
def procesar_archivo_invid(token):

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    data = tablaInvid(archivo=archivo)

    headers = {
        "Authorization": f"{token}"
    }

    # requests.post(url=url, json=data, headers=headers,
    #               timeout=None, verify=False)

    return make_response(jsonify(data), 200)


def tablaNb():
    # Abre el archivo CSV en modo lectura con la codificación adecuada
    with open(listadoCsv+"listadoNb.csv", "r", encoding="utf8") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=";")

        # Crea una lista para almacenar los datos
        data = []

        # Lee cada fila del archivo CSV (ignorando la primera fila de encabezados)

        next(csv_reader)  # Ignora la primera fila de encabezados
        for row in csv_reader:
            descripcion = row[3]
            categoria = row[2]
            precio = row[10]

            # Crea un diccionario con los datos de cada registro
            registro = {
                "proveedor": "nb",
                "producto": descripcion,
                "categoria": encontrar_valor(obtenerDiccionario("nb"), categoria),
                "precio": round((float(precio) * 1.1))
            }

            # Agrega el diccionario a la lista de datos
            data.append(registro)

    with open(listadoJson+"nb.json", 'w') as jf:
        json.dump(data, jf, ensure_ascii=False, indent=2)

    return data


@app.route("/procesar_archivo_nb", methods=["POST"])
@requiere_token
def procesar_archivo_nb(token):

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    # Guardar el archivo temporalmente
    archivo.save(listadoCsv+"listadoNb.csv")

    data = tablaNb()

    headers = {
        "Authorization": f"{token}"
    }

    # requests.post(url=url, json=data, headers=headers,
    #               timeout=None, verify=False)

    return make_response(jsonify(data), 200)


def tablaMega(archivo):
    df = pd.read_excel(archivo)

    # Se utilizan los índices 0, 1 y 2 para las primeras tres filas
    df = df.drop([0, 1])
    df.reset_index(drop=True, inplace=True)
    df.to_excel(listadosTemporales + "tempMega.xlsx", index=False)

    # Cargar el libro de trabajo
    book = load_workbook(listadosTemporales + "tempMega.xlsx")
    sheet = book["Sheet1"]

    def apply_custom_formula(value1, value2, value3, value4, value5, value6, value7):
        result = ""
        if (value1 is None or len(value1) <= 1) and (value2 is None or len(value2) <= 1):
            result = f"{value3} - {value4}"
        elif (value5 is None or len(value5) <= 1):
            result = ""
        elif (value2 is None or len(value2) <= 1):
            if (value6 is not None):
                posicion = value6.find(" - ")
                if posicion != -1:
                    parte_izquierda = value6[:posicion]
                    result = f"{parte_izquierda} - {value4}"
        else:
            result = f"{value7}"
        return result

    for row in range(3, sheet.max_row + 1):
        value1 = sheet[f'B{row-2}'].value
        value2 = sheet[f'B{row-1}'].value
        value3 = sheet[f'A{row-2}'].value
        value4 = sheet[f'A{row-1}'].value
        value5 = sheet[f'E{row}'].value
        value6 = sheet[f'G{row-2}'].value
        value7 = sheet[f'G{row-1}'].value
        result = apply_custom_formula(
            value1, value2, value3, value4, value5, value6, value7)

        cell = sheet.cell(row=row, column=7)  # Columna G
        cell.value = unidecode(result)

    # Save the changes to the file
    book.save(listadosTemporales + "tempMega.xlsx")

    # Guardar como CSV
    df = pd.read_excel(listadosTemporales+"tempMega.xlsx")
    df.to_csv(listadoCsv+"listadoMega.csv", index=False)

    # Abre el archivo CSV en modo lectura con la codificación adecuada
    with open(listadoCsv+"listadoMega.csv", "r", encoding="utf-8") as file:
        # Crea un objeto lector CSV
        csv_reader = csv.reader(file, delimiter=",")

        # Crea una lista para almacenar los datos
        data = []

        # Lee cada fila del archivo CSV (ignorando la primera fila de encabezados)

        next(csv_reader)  # Ignora la primera fila de encabezados
        for row in csv_reader:
            if (row[3] != ""):
                descripcion = row[1]
                precio = row[2].replace("U$s ", "")
                iva = row[4].replace("+", "").replace("%", "")
                categoria = row[6]

                # Crea un diccionario con los datos de cada registro
                registro = {
                    "proveedor": "mega",
                    "producto": descripcion,
                    "categoria": encontrar_valor(obtenerDiccionario("mega"), categoria),
                    "precio": round((float(precio) * (1 + (float(iva)/100)) * 1.1))
                }

                # Agrega el diccionario a la lista de datos

                data.append(registro)

    with open(listadoJson+"mega.json", 'w') as jf:
        json.dump(data, jf, ensure_ascii=False, indent=2)

    return data


@app.route("/procesar_archivo_mega", methods=["POST"])
@requiere_token
def procesar_archivo_mega(token):

    archivo = request.files["file"]

    if archivo.filename == "":
        return jsonify({"error": "No se ha seleccionado un archivo"}), 400

    # Guardar el archivo .xls en el directorio

    data = tablaMega(archivo=archivo)

    headers = {
        "Authorization": f"{token}"
    }

    # requests.post(url=url, json=data, headers=headers,
    #               timeout=None, verify=False)

    return make_response(jsonify(data), 200)


if __name__ == "__main__":
    certfile = 'secrets/fullchain.pem'
    keyfile = 'secrets/privkey.pem'

    import threading
    consumer_thread = threading.Thread(target=rabbitmq_consumer)
    consumer_thread.start()

    app.run(host='0.0.0.0', debug=True, ssl_context=(certfile, keyfile))
